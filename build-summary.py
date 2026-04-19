"""
Rebuilds the 'Summary' and 'Quarterly' tabs in Program-Evals-2026.xlsx with
live formulas. Auto-recalculates as new survey responses are pasted into raw
sheets — no Python re-run needed for routine updates.

Run when:
- Initial setup
- A new program sheet is added (TTTL3, etc.)
- Column layouts of existing sheets change
- You want to reset the boss's tabs from scratch

Creates a timestamped backup before modifying.
"""
import shutil
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = Path(r"C:\Users\Amy Miller - TS\OneDrive - TalentSmart\Program-Evals-2026.xlsx")
YEAR = 2026

# Per-sheet columns. "date" is the column we filter quarters on (Start Date).
SHEETS = {
    "TTTL1": {
        "label": "TTT L1", "family": "ttt",
        "date": "D", "respondent": "B", "modality": "S", "facilitator": "T",
        "content_relevant": "X", "fac_knowledge": "Y", "fac_engaged": "Z",
        "worthwhile": "AB", "apply_on_job": "AC", "gained_knowledge": "AD",
        "nps": "AG", "ei_dev_pct": "AJ", "confidence_pct": "AK",
        "manager_exp": None,
    },
    "TTTL2": {
        "label": "TTT L2", "family": "ttt",
        "date": "D", "respondent": "B", "modality": "R", "facilitator": "S",
        "content_relevant": "W", "fac_knowledge": "X", "fac_engaged": "Y",
        "worthwhile": "AA", "apply_on_job": "AB", "gained_knowledge": "AC",
        "nps": "AF", "ei_dev_pct": "AI", "confidence_pct": "AJ",
        "manager_exp": None,
    },
    "TTTTeams": {
        "label": "TTT Teams", "family": "ttt",
        "date": "D", "respondent": "B", "modality": "R", "facilitator": "S",
        "content_relevant": "W", "fac_knowledge": "X", "fac_engaged": "Y",
        "worthwhile": "AA", "apply_on_job": "AB", "gained_knowledge": "AC",
        "nps": "AF", "ei_dev_pct": "AI", "confidence_pct": "AJ",
        "manager_exp": None,
    },
    "PrivateL1": {
        "label": "Private L1", "family": "private",
        "date": "D", "respondent": "B", "modality": "O", "facilitator": "P",
        "content_relevant": "S", "fac_knowledge": "T", "fac_engaged": "U",
        "worthwhile": "W", "apply_on_job": "X", "gained_knowledge": "Y",
        "nps": "AB", "ei_dev_pct": "AD", "confidence_pct": "AE",
        "manager_exp": "AF",
    },
    "PrivateL2": {
        "label": "Private L2", "family": "private",
        "date": "D", "respondent": "B", "modality": "O", "facilitator": "P",
        "content_relevant": "S", "fac_knowledge": "T", "fac_engaged": "U",
        "worthwhile": "W", "apply_on_job": "X", "gained_knowledge": "Y",
        "nps": "AB", "ei_dev_pct": "AD", "confidence_pct": "AE",
        "manager_exp": "AF",
    },
    "PublicL1": {
        "label": "Public L1", "family": "public",
        "date": "D", "respondent": "B", "modality": "O", "facilitator": "P",
        "content_relevant": "S", "fac_knowledge": "T", "fac_engaged": "U",
        "worthwhile": "W", "apply_on_job": "X", "gained_knowledge": "Y",
        "nps": "AB", "ei_dev_pct": "AD", "confidence_pct": "AE",
        "manager_exp": None,
    },
    "Custom Programs": {
        "label": "Custom Programs", "family": "custom",
        "date": "D", "respondent": "B", "modality": "R", "facilitator": "S",
        "content_relevant": "V", "fac_knowledge": "W", "fac_engaged": "X",
        "worthwhile": "Z", "apply_on_job": "AA", "gained_knowledge": "AB",
        "nps": "AF", "ei_dev_pct": "AI", "confidence_pct": "AJ",
        "manager_exp": "AK",
    },
}

# Refresher is omitted from Summary/Quarterly because it uses different metrics.

QUARTERS = [
    ("Q1", (1, 1), (3, 31)),
    ("Q2", (4, 1), (6, 30)),
    ("Q3", (7, 1), (9, 30)),
    ("Q4", (10, 1), (12, 31)),
]

NAVY = "002D61"
GREEN = "0ACC8B"
GOLD = "DFA351"
BG_SOFT = "F7F8FB"
BG_SECTION = "EEF2F8"

LIKERT_KEYS = [
    ("apply_on_job", "% Will apply on the job (4–5)"),
    ("gained_knowledge", "% Gained new knowledge (4–5)"),
    ("worthwhile", "% Worthwhile investment (4–5)"),
    ("content_relevant", "% Content relevant (4–5)"),
    ("fac_knowledge", "% Facilitator knowledge enhanced learning (4–5)"),
    ("fac_engaged", "% Facilitator kept me engaged (4–5)"),
]


# ---------- Formula helpers ----------

def date_range(year, quarter_idx):
    _, (sm, sd), (em, ed) = QUARTERS[quarter_idx]
    return f"DATE({year},{sm},{sd})", f"DATE({year},{em},{ed})"

def q(sheet):
    """Quote a sheet name for use in formulas (handles spaces)."""
    return f"'{sheet}'" if " " in sheet else sheet

def date_filter(sheet, date_col, q_idx=None):
    """Returns the comma-separated COUNTIFS args for the date filter (or empty for YTD)."""
    if q_idx is None:
        return ""
    start, end = date_range(YEAR, q_idx)
    sn = q(sheet)
    return f',{sn}!{date_col}:{date_col},">="&{start},{sn}!{date_col}:{date_col},"<="&{end}'

def nps_formula_combined(sheets_subset, q_idx=None):
    proms, detrs, totals = [], [], []
    for sn, sc in sheets_subset.items():
        d, n = sc["date"], sc["nps"]
        df = date_filter(sn, d, q_idx)
        sq = q(sn)
        proms.append(f'COUNTIFS({sq}!{n}:{n},">=9"{df})')
        detrs.append(f'COUNTIFS({sq}!{n}:{n},"<=6"{df})')
        totals.append(f'COUNTIFS({sq}!{n}:{n},">=0"{df})')
    p, dn, t = "+".join(proms), "+".join(detrs), "+".join(totals)
    return f'=IFERROR(ROUND(({p}-({dn}))/({t})*100,0),"-")'

def t2b_formula_combined(sheets_subset, key, q_idx=None):
    tops, totals = [], []
    for sn, sc in sheets_subset.items():
        col = sc[key]
        d = sc["date"]
        df = date_filter(sn, d, q_idx)
        sq = q(sn)
        tops.append(f'COUNTIFS({sq}!{col}:{col},">=4"{df})')
        totals.append(f'COUNTIFS({sq}!{col}:{col},">=0"{df})')
    return f'=IFERROR(ROUND(({"+".join(tops)})/({"+".join(totals)})*100,0),"-")'

def participants_formula_combined(sheets_subset, q_idx=None):
    parts = []
    for sn, sc in sheets_subset.items():
        d, r = sc["date"], sc["respondent"]
        df = date_filter(sn, d, q_idx)
        sq = q(sn)
        parts.append(f'COUNTIFS({sq}!{r}:{r},"<>"{df})')
    return f'={"+".join(parts)}'

def avg_pct_formula_combined(sheets_subset, key, q_idx=None):
    sums, counts = [], []
    for sn, sc in sheets_subset.items():
        col = sc[key]
        d = sc["date"]
        sq = q(sn)
        if q_idx is None:
            sums.append(f'SUM({sq}!{col}:{col})')
            counts.append(f'COUNT({sq}!{col}:{col})')
        else:
            start, end = date_range(YEAR, q_idx)
            sums.append(f'SUMIFS({sq}!{col}:{col},'
                        f'{sq}!{d}:{d},">="&{start},'
                        f'{sq}!{d}:{d},"<="&{end})')
            counts.append(f'COUNTIFS({sq}!{col}:{col},">=0",'
                          f'{sq}!{d}:{d},">="&{start},'
                          f'{sq}!{d}:{d},"<="&{end})')
    return f'=IFERROR(ROUND(({"+".join(sums)})/({"+".join(counts)}),0),"-")'

def no_manager_formula(sheets_with_q, q_idx=None):
    """Sheets must have manager_exp column. Returns % NO."""
    nos, totals = [], []
    for sn, sc in sheets_with_q.items():
        col = sc["manager_exp"]
        if not col: continue
        d = sc["date"]
        df = date_filter(sn, d, q_idx)
        sq = q(sn)
        nos.append(f'COUNTIFS({sq}!{col}:{col},"No"{df})')
        totals.append(f'(COUNTIFS({sq}!{col}:{col},"No"{df})+COUNTIFS({sq}!{col}:{col},"Yes"{df}))')
    if not nos:
        return '"-"'
    return f'=IFERROR(ROUND(({"+".join(nos)})/({"+".join(totals)})*100,0),"-")'

def sessions_distinct_formula(sheets_subset):
    """Distinct session names across the given sheets, using SUMPRODUCT for compatibility."""
    parts = []
    for sn, sc in sheets_subset.items():
        sq = q(sn)
        parts.append(f'SUMPRODUCT(({sq}!A2:A1000<>"")/COUNTIF({sq}!A2:A1000,{sq}!A2:A1000&""))')
    return f'=IFERROR(ROUND({"+".join(parts)},0),0)'

def modality_count_formula(sheets_subset, value):
    parts = []
    for sn, sc in sheets_subset.items():
        sq = q(sn)
        parts.append(f'COUNTIF({sq}!{sc["modality"]}:{sc["modality"]},"{value}")')
    return f'={"+".join(parts)}'


# ---------- Tab builders ----------

def style_title(ws, cell_ref, text, span_cols):
    last_col = chr(ord(cell_ref[0]) + span_cols - 1)
    ws.merge_cells(f"{cell_ref}:{last_col}{cell_ref[1:]}")
    cell = ws[cell_ref]
    cell.value = text
    cell.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[int(cell_ref[1:])].height = 36

def style_subtitle(ws, cell_ref, text, span_cols):
    last_col = chr(ord(cell_ref[0]) + span_cols - 1)
    ws.merge_cells(f"{cell_ref}:{last_col}{cell_ref[1:]}")
    cell = ws[cell_ref]
    cell.value = text
    cell.font = Font(name="Calibri", size=10, italic=True, color="555555")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[int(cell_ref[1:])].height = 22

def style_section(ws, row, text, ncols):
    ws.cell(row=row, column=1, value=text).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BG_SECTION)
    ws.row_dimensions[row].height = 22

def style_header(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


def build_summary_tab(wb):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    # Programs across the top: each individual program + family summaries + All
    program_cols = [
        ("All Programs", SHEETS),
        ("TTT Summary", {k: v for k, v in SHEETS.items() if v["family"] == "ttt"}),
        ("TTT L1", {"TTTL1": SHEETS["TTTL1"]}),
        ("TTT L2", {"TTTL2": SHEETS["TTTL2"]}),
        ("TTT Teams", {"TTTTeams": SHEETS["TTTTeams"]}),
        ("Private Summary", {k: v for k, v in SHEETS.items() if v["family"] == "private"}),
        ("Private L1", {"PrivateL1": SHEETS["PrivateL1"]}),
        ("Private L2", {"PrivateL2": SHEETS["PrivateL2"]}),
        ("Public L1", {"PublicL1": SHEETS["PublicL1"]}),
        ("Custom Programs", {"Custom Programs": SHEETS["Custom Programs"]}),
    ]
    ncols = 1 + len(program_cols)

    # Column widths
    ws.column_dimensions["A"].width = 44
    for i in range(2, ncols + 1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = 14

    style_title(ws, "A1", "Program Impact — Summary (Year to Date)", ncols)
    style_subtitle(ws, "A2", "Auto-calculated from the raw response sheets. Refresh by adding rows there.", ncols)

    header_row = 4
    style_header(ws, header_row, ["Metric"] + [p[0] for p in program_cols])

    r = 5
    style_section(ws, r, "OVERVIEW", ncols)
    r += 1
    ws.cell(row=r, column=1, value="Respondents (EOS completed)")
    for i, (_, sheets_subset) in enumerate(program_cols, start=2):
        ws.cell(row=r, column=i, value=participants_formula_combined(sheets_subset)).alignment = Alignment(horizontal="center")
    r += 1
    ws.cell(row=r, column=1, value="Sessions delivered (distinct)")
    for i, (_, sheets_subset) in enumerate(program_cols, start=2):
        ws.cell(row=r, column=i, value=sessions_distinct_formula(sheets_subset)).alignment = Alignment(horizontal="center")
    r += 1
    ws.cell(row=r, column=1, value="Virtual respondents")
    for i, (_, sheets_subset) in enumerate(program_cols, start=2):
        ws.cell(row=r, column=i, value=modality_count_formula(sheets_subset, "Virtual")).alignment = Alignment(horizontal="center")
    r += 1
    ws.cell(row=r, column=1, value="In-person respondents")
    for i, (_, sheets_subset) in enumerate(program_cols, start=2):
        ws.cell(row=r, column=i, value=modality_count_formula(sheets_subset, "In person")).alignment = Alignment(horizontal="center")
    r += 2

    style_section(ws, r, "NPS", ncols); r += 1
    ws.cell(row=r, column=1, value="Net Promoter Score")
    for i, (_, sheets_subset) in enumerate(program_cols, start=2):
        ws.cell(row=r, column=i, value=nps_formula_combined(sheets_subset)).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=i).font = Font(bold=True)
    r += 2

    style_section(ws, r, "TOP-2-BOX % (rated 4 or 5)", ncols); r += 1
    for key, label in LIKERT_KEYS:
        ws.cell(row=r, column=1, value=label)
        for i, (_, sheets_subset) in enumerate(program_cols, start=2):
            ws.cell(row=r, column=i, value=t2b_formula_combined(sheets_subset, key)).alignment = Alignment(horizontal="center")
        r += 1
    r += 1

    style_section(ws, r, "EI DEVELOPMENT ATTRIBUTION", ncols); r += 1
    ws.cell(row=r, column=1, value="Avg % EI growth credited to training")
    for i, (_, sheets_subset) in enumerate(program_cols, start=2):
        ws.cell(row=r, column=i, value=avg_pct_formula_combined(sheets_subset, "ei_dev_pct")).alignment = Alignment(horizontal="center")
    r += 1
    ws.cell(row=r, column=1, value="Avg confidence in estimate")
    for i, (_, sheets_subset) in enumerate(program_cols, start=2):
        ws.cell(row=r, column=i, value=avg_pct_formula_combined(sheets_subset, "confidence_pct")).alignment = Alignment(horizontal="center")
    r += 2

    style_section(ws, r, "MANAGER EXPECTATIONS (where asked)", ncols); r += 1
    ws.cell(row=r, column=1, value="% who said NO (manager did not set expectations)")
    for i, (_, sheets_subset) in enumerate(program_cols, start=2):
        applicable = {k: v for k, v in sheets_subset.items() if v.get("manager_exp")}
        if applicable:
            ws.cell(row=r, column=i, value=no_manager_formula(applicable)).alignment = Alignment(horizontal="center")
        else:
            ws.cell(row=r, column=i, value="n/a").alignment = Alignment(horizontal="center")

    ws.freeze_panes = "B5"


def build_quarterly_tab(wb):
    if "Quarterly" in wb.sheetnames:
        del wb["Quarterly"]
    ws = wb.create_sheet("Quarterly", 1)
    ws.sheet_view.showGridLines = False

    cols = ["Metric", "Q1", "Q2", "Q3", "Q4", "YTD"]
    ws.column_dimensions["A"].width = 50
    for i, c in enumerate("BCDEF", start=2):
        ws.column_dimensions[c].width = 12

    style_title(ws, "A1", f"Quarterly Breakdown — {YEAR}", 6)
    style_subtitle(ws, "A2", "All numbers auto-recalculate from the raw response sheets. Quarter is determined by Start Date.", 6)

    style_header(ws, 4, cols)

    program_groups = [
        ("All Programs", SHEETS),
        ("TTT Summary", {k: v for k, v in SHEETS.items() if v["family"] == "ttt"}),
        ("TTT L1", {"TTTL1": SHEETS["TTTL1"]}),
        ("TTT L2", {"TTTL2": SHEETS["TTTL2"]}),
        ("TTT Teams", {"TTTTeams": SHEETS["TTTTeams"]}),
        ("Private Summary", {k: v for k, v in SHEETS.items() if v["family"] == "private"}),
        ("Private L1", {"PrivateL1": SHEETS["PrivateL1"]}),
        ("Private L2", {"PrivateL2": SHEETS["PrivateL2"]}),
        ("Public L1", {"PublicL1": SHEETS["PublicL1"]}),
        ("Custom Programs", {"Custom Programs": SHEETS["Custom Programs"]}),
    ]

    r = 5
    for prog_label, subset in program_groups:
        style_section(ws, r, prog_label.upper(), 6); r += 1

        def write_row(label, formula_fn):
            ws.cell(row=r, column=1, value=label)
            for q in range(4):
                cell = ws.cell(row=r, column=2 + q, value=formula_fn(subset, q_idx=q))
                cell.alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=6, value=formula_fn(subset, q_idx=None)).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=6).font = Font(bold=True)

        write_row("Participants", participants_formula_combined)
        r += 1
        write_row("NPS", nps_formula_combined)
        r += 1
        for key, label in LIKERT_KEYS:
            ws.cell(row=r, column=1, value=label)
            for q in range(4):
                ws.cell(row=r, column=2 + q, value=t2b_formula_combined(subset, key, q_idx=q)).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=6, value=t2b_formula_combined(subset, key, q_idx=None)).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=6).font = Font(bold=True)
            r += 1

        # manager exp where applicable
        applicable = {k: v for k, v in subset.items() if v.get("manager_exp")}
        if applicable:
            ws.cell(row=r, column=1, value="% who said NO their manager set expectations")
            for q in range(4):
                ws.cell(row=r, column=2 + q, value=no_manager_formula(applicable, q_idx=q)).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=6, value=no_manager_formula(applicable, q_idx=None)).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=6).font = Font(bold=True)
            r += 1
        r += 1  # blank row between programs

    ws.freeze_panes = "B5"


def main():
    if not SRC.exists():
        raise SystemExit(f"Source file not found: {SRC}")
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = SRC.with_name(SRC.stem + f"-backup-{ts}.xlsx")
    shutil.copy2(SRC, backup)
    print(f"Backup created: {backup.name}")

    wb = openpyxl.load_workbook(SRC)
    build_summary_tab(wb)
    build_quarterly_tab(wb)
    wb.save(SRC)
    print(f"Summary + Quarterly tabs rebuilt in: {SRC.name}")


if __name__ == "__main__":
    main()

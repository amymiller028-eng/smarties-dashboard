"""
Reads Program-Evals-2026.xlsx and writes the dashboard's data.json.

Run any time you've added new survey responses to the workbook.

    python compute-data.py
"""
import json
import re
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl

SRC = Path(r"C:\Users\Amy Miller - TS\OneDrive - TalentSmart\Program-Evals-2026.xlsx")
OUT = Path(__file__).parent / "data.json"

# 2025 lives in its own workbook and is closed history — read, never written.
# Its sheets ask the same questions in different columns, which resolve_cols()
# handles, so each one just names the 2026 sheet whose history it continues.
LEGACY_SRC = Path(r"C:\Users\Amy Miller - TS\OneDrive - TalentSmart\Program_Evals.xlsx")
LEGACY_SHEETS = {
    "PrivateL1_EOS": "PrivateL1",
    "PrivateL2_EOS": "PrivateL2",
    "TTTLevel1_EOS": "TTTL1",
    "TTTLevel2_EOS": "TTTL2",
    "TTT_Teams":     "TTTTeams",
    # EQ in Policing has no 2026 counterpart — it becomes its own family under
    # Custom Programs. NOTE: the sheet holds 3 sessions, only 19 of its 26 rows
    # are the policing engagement itself.
    "policing":      "Policing",
    # "Refresh" is deliberately absent: despite the name it's a standard EOS
    # survey, not the refresher-readiness instrument, and its item labels sit
    # in row 2. It needs its own parse before it can be included.
}

# --- Column resolution --------------------------------------------------------
# Columns are found by HEADER TEXT and then confirmed by the SHAPE OF THEIR
# VALUES, because a header row can't be trusted on its own: the TTTTeams export
# has a header row sitting one column to the right of its own data, so matching
# on text alone reads the wrong column for every metric on that sheet. Anything
# in COLUMN_OVERRIDES below wins outright.

COLUMN_NEEDLES = {
    "session":          ["session name"],
    "modality":         ["facilitated virtually"],
    "facilitator":      ["facilitator name"],
    "content_relevant": ["relevant to my job"],
    "fac_knowledge":    ["enhanced by my facilitator"],
    "fac_engaged":      ["facilitator kept me engaged"],
    "worthwhile":       ["worthwhile investment"],
    "apply_on_job":     ["apply what i learned"],
    "gained_knowledge": ["gained new knowledge"],
    "nps":              ["recommend this program"],
    "ei_dev_pct":       ["percentage of your development in emotional intelligence"],
    "confidence_pct":   ["confident are you in this estimate"],
    "manager_exp":      ["manager communicated"],
    "quote":            ["share with future participants", "share with future facilitators"],
    "start_date":       ["start date"],
    "end_date":         ["end date"],
}


def _nums(values):
    return [v for v in values if isinstance(v, (int, float))]

def _in_range(lo, hi):
    def check(values):
        n = _nums(values)
        return bool(n) and all(lo <= v <= hi for v in n)
    return check

def _looks_like_text(values):
    return any(isinstance(v, str) and v.strip() and not v.strip().isdigit()
               for v in values)

def _looks_like_modality(values):
    return any(str(v).strip().lower() in ("virtual", "in person")
               for v in values if v not in (None, ""))

_LIKERT = _in_range(1, 5)

# A detected column must also look like its metric, or we refuse it.
COLUMN_SHAPES = {
    "modality":         _looks_like_modality,
    "facilitator":      _looks_like_text,
    "session":          _looks_like_text,
    "quote":            _looks_like_text,
    "nps":              _in_range(0, 10),
    "ei_dev_pct":       _in_range(0, 100),
    "confidence_pct":   _in_range(0, 100),
    "content_relevant": _LIKERT,
    "fac_knowledge":    _LIKERT,
    "fac_engaged":      _LIKERT,
    "worthwhile":       _LIKERT,
    "apply_on_job":     _LIKERT,
    "gained_knowledge": _LIKERT,
}

# Sheets whose headers can't be trusted, or questions to deliberately ignore.
# None means "this sheet doesn't have that question, don't go looking".
COLUMN_OVERRIDES = {
    # Header row is one column right of the data — every key pinned by hand.
    "TTTTeams": {
        "session": "A", "modality": "R", "facilitator": "S",
        "content_relevant": "W", "fac_knowledge": "X", "fac_engaged": "Y",
        "worthwhile": "AA", "apply_on_job": "AB", "gained_knowledge": "AC",
        "nps": "AF", "ei_dev_pct": "AI", "confidence_pct": "AJ",
        "manager_exp": None, "quote": "AK",
    },
}


# Legacy hand-written map. Kept as the reference the resolver is checked
# against; `--verify-columns` reports any disagreement.
SHEET_COLS = {
    "TTTL1": {
        "session": "A", "modality": "S", "facilitator": "T",
        "content_relevant": "X", "fac_knowledge": "Y", "fac_engaged": "Z",
        "worthwhile": "AB", "apply_on_job": "AC", "gained_knowledge": "AD",
        "nps": "AG", "ei_dev_pct": "AJ", "confidence_pct": "AK",
        "manager_exp": None, "quote": "AL",
    },
    "TTTTeams": {
        # NOTE: header row in this sheet is misaligned with data — data follows
        # the TTTL2 layout (one column to the left of TTTL1 headers).
        "session": "A", "modality": "R", "facilitator": "S",
        "content_relevant": "W", "fac_knowledge": "X", "fac_engaged": "Y",
        "worthwhile": "AA", "apply_on_job": "AB", "gained_knowledge": "AC",
        "nps": "AF", "ei_dev_pct": "AI", "confidence_pct": "AJ",
        "manager_exp": None, "quote": "AK",
    },
    "TTTL2": {
        "session": "A", "modality": "R", "facilitator": "S",
        "content_relevant": "W", "fac_knowledge": "X", "fac_engaged": "Y",
        "worthwhile": "AA", "apply_on_job": "AB", "gained_knowledge": "AC",
        "nps": "AF", "ei_dev_pct": "AI", "confidence_pct": "AJ",
        "manager_exp": None, "quote": "AK",
    },
    "PrivateL1": {
        "session": "A", "modality": "O", "facilitator": "P",
        "content_relevant": "S", "fac_knowledge": "T", "fac_engaged": "U",
        "worthwhile": "W", "apply_on_job": "X", "gained_knowledge": "Y",
        "nps": "AB", "ei_dev_pct": "AD", "confidence_pct": "AE",
        "manager_exp": "AF", "quote": "AG",
    },
    "PrivateL2": {
        "session": "A", "modality": "O", "facilitator": "P",
        "content_relevant": "S", "fac_knowledge": "T", "fac_engaged": "U",
        "worthwhile": "W", "apply_on_job": "X", "gained_knowledge": "Y",
        "nps": "AB", "ei_dev_pct": "AD", "confidence_pct": "AE",
        "manager_exp": "AF", "quote": "AG",
    },
    "PublicL1": {
        "session": "A", "modality": "O", "facilitator": "P",
        "content_relevant": "S", "fac_knowledge": "T", "fac_engaged": "U",
        "worthwhile": "W", "apply_on_job": "X", "gained_knowledge": "Y",
        "nps": "AB", "ei_dev_pct": "AD", "confidence_pct": "AE",
        "manager_exp": None, "quote": "AG",
    },
    "Custom Programs": {
        "session": "A", "modality": "R", "facilitator": "S",
        "content_relevant": "V", "fac_knowledge": "W", "fac_engaged": "X",
        "worthwhile": "Z", "apply_on_job": "AA", "gained_knowledge": "AB",
        "nps": "AF", "ei_dev_pct": "AI", "confidence_pct": "AJ",
        "manager_exp": "AK", "quote": "AL",
    },
}

# --- Leading Through Friction -------------------------------------------------
# LTF runs its own survey: seven agreement statements that only partly overlap
# the standard question set. One sheet per delivery mode, named "LTF-<Delivery>".
# Add LTF-Public / LTF-TTT sheets and they appear on the dashboard by themselves.
LTF_DELIVERIES = {                       # sheet suffix -> (view key, tab label)
    "Private": ("ltf-private", "Private Program"),
    "Public":  ("ltf-public",  "Public"),
    "TTT":     ("ltf-ttt",     "Train the Trainer"),
}

# LTF's own columns, named for the questions they actually ask.
LTF_COLS = {
    "session": "A", "modality": "L", "facilitator": "M",
    "materials": "N", "engaged": "O", "tools": "P", "worthwhile": "Q",
    "clearer": "R", "improve": "S", "recommend_team": "T",
    "nps": "W", "learned": "X",
}

# The seven statements, ordered strongest-first. Drives the dashboard list.
LTF_STATEMENTS = [
    ("tools",          "I gained practical tools I can apply to real leadership challenges"),
    ("worthwhile",     "This course was a worthwhile investment of my time"),
    ("recommend_team", "I'd want my team or other leaders in my organization to go through this"),
    ("engaged",        "The facilitator kept me engaged"),
    ("materials",      "The course materials were easy to navigate"),
    ("clearer",        "I have a clearer way to recognize and respond to friction when it shows up"),
    ("improve",        "I expect this program to improve how I lead going forward"),
]

# How LTF feeds the SHARED "All Programs" tiles. Only four of the six have an
# honest equivalent: LTF never asks about content relevance or facilitator
# knowledge, so it contributes nothing to those two rather than letting a
# near-enough question stand in for them.
LTF_AS_STANDARD = {
    "session": "A", "modality": "L", "facilitator": "M",
    "content_relevant": None,        # not asked in LTF
    "fac_knowledge": None,           # not asked in LTF
    "fac_engaged": "O",              # "The facilitator kept me engaged"
    "worthwhile": "Q",               # exact match
    "apply_on_job": "P",             # "gained practical tools I can apply..."
    "gained_knowledge": "R",         # "clearer way to recognize and respond to friction"
    "nps": "W",
    "ei_dev_pct": None,              # LTF doesn't ask the EQ-attribution pair,
    "confidence_pct": None,          # so it can't drag those averages either way
    "manager_exp": None,
    "quote": "X",                    # "most important thing you learned"
}


# --- Dates --------------------------------------------------------------------
# Response dates arrive in four states. Real datetimes; Excel serial numbers
# saved as text ("45891.59349537037"); the literal string "########", which is
# Excel's too-narrow-column display accidentally saved as a value and destroys
# the date in that cell; and blank. For the last two the date is recovered from
# the Session Name, which nearly always carries one ("... - July 23, 2026").

EXCEL_EPOCH = datetime(1899, 12, 30)

_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"], start=1)}

_SESSION_MONTH_YEAR = re.compile(
    r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s*"
    r"(?:\d{1,2}(?:\s*[-–]\s*\d{1,2})?\s*,?\s*)?(20\d{2})", re.I)
_SESSION_YEAR_ONLY = re.compile(r"\b(20\d{2})\b")
_SESSION_YEAR_RANGE = re.compile(r"\b20\d{2}\s*[-–]\s*20\d{2}\b")


def parse_cell_date(value):
    """A datetime from a date cell, whatever state it's in. None if unusable."""
    if isinstance(value, datetime):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text or set(text) == {"#"}:          # '########' — destroyed
        return None
    try:                                        # Excel serial saved as text
        return EXCEL_EPOCH + timedelta(days=float(text))
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d%b%Y", "%B %d, %Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_session_date(session):
    """Approximate date from a session name — month precision is enough to
    bucket by year and quarter. None when the name spans years ("2023-2024")."""
    if not session:
        return None
    text = " ".join(str(session).split())
    if _SESSION_YEAR_RANGE.search(text):        # ambiguous, don't guess
        return None
    m = _SESSION_MONTH_YEAR.search(text)
    if m:
        return datetime(int(m.group(2)), _MONTHS[m.group(1).lower()[:3]], 1)
    y = _SESSION_YEAR_ONLY.search(text)
    return datetime(int(y.group(1)), 1, 1) if y else None


def response_date(start, end, session):
    """Best available date for one response, most trustworthy source first."""
    return (parse_cell_date(start) or parse_cell_date(end)
            or parse_session_date(session))


def period_of(dt):
    """{'year': 2026, 'quarter': '2026-Q3'} — or None when undated."""
    return None if dt is None else {
        "year": dt.year, "quarter": f"{dt.year}-Q{(dt.month - 1) // 3 + 1}"}


def _read_locked_file(path):
    """Read a file Excel currently has open.

    Python's open() asks Windows for a handle in a way Excel's lock refuses,
    even though the file is perfectly readable — Explorer and PowerShell copy
    it fine. Going through CreateFileW with every share flag set gets the same
    access they do.
    """
    import ctypes, msvcrt, os
    from ctypes import wintypes

    GENERIC_READ, OPEN_EXISTING, FILE_ATTRIBUTE_NORMAL = 0x80000000, 3, 0x80
    FILE_SHARE_ALL = 0x1 | 0x2 | 0x4          # read | write | delete

    CreateFileW = ctypes.windll.kernel32.CreateFileW
    CreateFileW.argtypes = [wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD,
                            wintypes.LPVOID, wintypes.DWORD, wintypes.DWORD,
                            wintypes.HANDLE]
    CreateFileW.restype = wintypes.HANDLE

    handle = CreateFileW(str(path), GENERIC_READ, FILE_SHARE_ALL, None,
                         OPEN_EXISTING, FILE_ATTRIBUTE_NORMAL, None)
    if handle == wintypes.HANDLE(-1).value:
        raise ctypes.WinError(ctypes.get_last_error())
    with os.fdopen(msvcrt.open_osfhandle(handle, os.O_RDONLY), "rb") as fh:
        return fh.read()


def load_readonly(path, **kwargs):
    """Open a workbook we only ever read. Having it open in Excel shouldn't
    block a refresh, so fall back to reading a snapshot of the bytes."""
    try:
        return openpyxl.load_workbook(path, **kwargs)
    except PermissionError:
        import io
        try:
            data = _read_locked_file(path)
        except Exception as exc:
            print(f"  ! {path.name} is locked and could not be read ({exc}). "
                  f"Close it in Excel and re-run.")
            raise
        print(f"  ({path.name} is open in Excel — reading a snapshot of it)")
        return openpyxl.load_workbook(io.BytesIO(data), **kwargs)


def _header_map(ws):
    """{'A': 'session name', ...} — normalised header text by column letter."""
    from openpyxl.utils import get_column_letter
    out = {}
    for c in range(1, min(ws.max_column, 120) + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            out[get_column_letter(c)] = " ".join(str(h).split()).lower()
    return out


def resolve_cols(ws, sheet_name, needles=None, shapes=None, warn=None):
    """Locate each metric's column: explicit override first, else header text
    confirmed against the shape of the values underneath it.

    Returns {key: 'G' or None}. Appends human-readable notes to `warn`.
    """
    needles = needles or COLUMN_NEEDLES
    shapes = shapes or COLUMN_SHAPES
    warn = warn if warn is not None else []
    overrides = COLUMN_OVERRIDES.get(sheet_name, {})
    headers = _header_map(ws)

    resolved = {}
    for key in needles:
        if key in overrides:                       # pinned by hand — trust it
            resolved[key] = overrides[key]
            continue

        letter = next((col for col, text in headers.items()
                       if any(n in text for n in needles[key])), None)
        if letter is None:
            resolved[key] = None
            continue

        values = col_values(ws, letter)
        if not any(v not in (None, "") for v in values):
            resolved[key] = None                   # column exists but is empty
            continue

        check = shapes.get(key)
        if check and not check(values):
            warn.append(f"{sheet_name}.{key}: header matched at {letter} but the "
                        f"values don't look like {key} — ignoring this column")
            resolved[key] = None
            continue

        resolved[key] = letter
    return resolved


# --- 90-day Learner impact survey ---------------------------------------------
# A different instrument entirely: sent ~90 days after a program to measure what
# people still DO, not how they felt on the day. Lives in the 2025 workbook and
# keeps collecting, so it's read by cohort rather than by year.
LEARNER_SHEET = "Learner"

LEARNER_SCALE = {"strongly disagree": 1, "disagree": 2,
                 "neither agree or disagree": 3, "neither agree nor disagree": 3,
                 "agree": 4, "strongly agree": 5}

# Items where AGREEING is the bad outcome — improvement means the score falls.
# Scored naively these read as declines and drag the headline down.
LEARNER_REVERSED = ["hard time trusting", "burned out", "do not handle change"]

# Seven job levels is too thin to slice once you account for the 67-of-91 who
# completed the before/after block, so they're banded into three.
LEARNER_BANDS = [
    ("Individual contributor", ["employee/associate", "supervisor",
                                "independently employed/consultant"]),
    ("Manager",                ["manager/senior manager"]),
    ("Senior leader",          ["director/senior director", "executive/vp",
                                "senior executive/svp"]),
]

# Multi-select blocks: (label, first column header, last column header).
# Bounds are pinned by hand — inferring them swallowed the neighbouring matrix.
LEARNER_BLOCKS = [
    ("Business outcomes improved",
     "please select any measurable business outcomes", "other (please specify)"),
    ("Where it helped most",
     "in which area(s) has the training helped", "other"),
    ("Barriers to applying it",
     "most significant barriers", "other (please specify)4"),
]

LEARNER_SINGLES = {
    "nps":         "how likely is it that you would recommend",
    "overall":     "how would you rate the training experience",
    "applies":     "how often do you apply the skills",
    "challenges":  "how often have you encountered any challenges",
    "manager":     "extent has your manager talked with you",
    "job_level":   "what is your job level",
    "session":     "session",
    "ei_pct":      "what percentage of your development in emotional intelligence",
}


def ltf_sheets(sheetnames):
    """LTF delivery sheets present in the workbook, in delivery order."""
    return [f"LTF-{sfx}" for sfx in LTF_DELIVERIES if f"LTF-{sfx}" in sheetnames]


def all_sheet_cols(wb, warn=None):
    """Standard-shaped column map for every sheet, LTF sheets included.

    Resolved from the workbook itself rather than hardcoded, so a sheet whose
    columns shift (or a new year's export) maps itself.
    """
    cols = {}
    for name in SHEET_COLS:
        if name in wb.sheetnames:
            cols[name] = resolve_cols(wb[name], name, warn=warn)
    for name in ltf_sheets(wb.sheetnames):
        cols[name] = LTF_AS_STANDARD
    return cols


CONFIDENCE_MAP = {
    "slightly confident": 1,
    "moderately confident": 2,
    "confident": 3,
    "fully confident": 4,
}
VALUABLE_TOP = {"very valuable", "extremely valuable"}

# --- Quote attribution (participant name + company) ---------------------------
# How names render on quotes: "first" → "Jane", "first_initial" → "Jane S.",
# "full" → "Jane Smith". On a public site "first" keeps exposure low; switch
# this one value to show fuller names.
NAME_STYLE = "first"

# Email domains that are personal, not a company — never shown as a "company" tag.
GENERIC_EMAIL_DOMAINS = {
    "gmail", "googlemail", "yahoo", "ymail", "rocketmail", "hotmail", "outlook",
    "live", "msn", "icloud", "me", "mac", "aol", "aim", "proton", "protonmail",
    "gmx", "mail", "zoho", "yandex", "comcast", "verizon", "att", "sbcglobal",
    "bellsouth", "cox", "charter",
}

# Header text used to locate the name/email/strengths columns per sheet.
# Detected by header (not fixed letters) so it survives the per-sheet column
# shifts. First match wins, so list more-specific needles first.
EXTRA_HEADER_NEEDLES = {
    "email": ["email address", "email"],
    "first": ["first name"],
    "last": ["last name"],
    "name_optional": ["name (optional)"],
    "strengths": ["facilitator's strengths", "facilitator strengths", "strengths"],
}


def col_values(ws, col):
    return [ws[f"{col}{r}"].value for r in range(2, ws.max_row + 1)]

def numeric(values):
    out = []
    for v in values:
        if v is None or v == "":
            continue
        try:
            out.append(float(v))
        except (TypeError, ValueError):
            pass
    return out

def nps(scores):
    nums = numeric(scores)
    if not nums:
        return 0
    promoters = sum(1 for s in nums if s >= 9)
    detractors = sum(1 for s in nums if s <= 6)
    return round((promoters - detractors) / len(nums) * 100)

def top2box(scores):
    nums = numeric(scores)
    if not nums:
        return 0
    return round(sum(1 for s in nums if s >= 4) / len(nums) * 100)

def avg(values):
    nums = numeric(values)
    return round(sum(nums) / len(nums)) if nums else 0

def count_text(values, target):
    return sum(1 for v in values if v and str(v).strip().lower() == target.lower())


# --- Categorical answers -------------------------------------------------------
# 2026 exports store the answer text ("Yes", "In Person"). The 2025 exports often
# store the option NUMBER instead, and sometimes a longer label. Matching on the
# literal string silently counted those as neither, which understated the
# manager-expectations figure badly (3% where the real answer is ~64%).

def norm_yes_no(value):
    """'1'/'Yes' -> 'yes', '2'/'No' -> 'no', anything else -> '' ."""
    t = str(value).strip().lower()
    if t in ("1", "yes", "y"):
        return "yes"
    if t in ("2", "no", "n"):
        return "no"
    return ""


def norm_modality(value):
    """'Virtual (video-based, online) Delivery' -> 'virtual', 'In Person' ->
    'in person'. The 2025 exports store the option number instead: 1 = Virtual,
    2 = In Person (confirmed by Amy, 2026-07-30), matching how 1/2 encode
    Yes/No for the manager-expectations question."""
    t = str(value).strip().lower()
    if t.startswith("virtual"):
        return "virtual"
    if t.startswith("in person") or t.startswith("in-person"):
        return "in person"
    if t == "1":
        return "virtual"
    if t == "2":
        return "in person"
    return ""


def count_norm(values, normaliser, target):
    return sum(1 for v in values if v not in (None, "") and normaliser(v) == target)

def distinct_count(values):
    return len({v for v in values if v not in (None, "")})


def collect_sheet(ws, cols):
    """Pull all relevant column values from a sheet at once, plus a per-row
    response date parallel to them (`_date`) so views can be sliced by period."""
    ds = {key: col_values(ws, c) if c else [] for key, c in cols.items()}

    n = max((len(v) for v in ds.values()), default=0)
    def at(key, i):
        col = ds.get(key) or []
        return col[i] if i < len(col) else None

    ds["_date"] = [response_date(at("start_date", i), at("end_date", i),
                                 at("session", i)) for i in range(n)]
    return ds


def filter_period(ds, period):
    """A copy of one dataset holding only responses from `period` ('2025',
    '2026', or 'all'). Undated responses are kept only in 'all'."""
    if period == "all":
        return ds
    year = int(period)
    keep = [i for i, d in enumerate(ds.get("_date", [])) if d and d.year == year]
    out = {}
    for key, col in ds.items():
        out[key] = [col[i] for i in keep if i < len(col)] if isinstance(col, list) else col
    return out


def periods_in(datasets):
    """Sorted list of years present across datasets."""
    years = {d.year for ds in datasets for d in ds.get("_date", []) if d}
    return [str(y) for y in sorted(years)]


def standard_view(label, datasets):
    """Aggregate one or more dataset dicts into a standard view object."""
    def cat(key):
        out = []
        for ds in datasets:
            out.extend(ds.get(key, []))
        return out

    sessions_all = [s for s in cat("session") if s]
    nps_all = cat("nps")
    participants = sum(len(numeric(ds.get("nps", []))) for ds in datasets)
    if participants == 0:
        # fall back to count of rows that have anything in modality
        participants = sum(sum(1 for v in ds.get("modality", []) if v) for ds in datasets)

    view = {
        "label": label,
        "type": "standard",
        "nps": nps(nps_all),
        "participants": participants,
        "sessions": distinct_count(sessions_all),
        "clients": distinct_count(sessions_all),  # 1 client per session as proxy; editable
        "eiDevelopmentAttributed": avg(cat("ei_dev_pct")),
        "confidenceInEstimate": avg(cat("confidence_pct")),
        "topBox": {
            "applyOnJob": top2box(cat("apply_on_job")),
            "gainedKnowledge": top2box(cat("gained_knowledge")),
            "worthwhileInvestment": top2box(cat("worthwhile")),
            "contentRelevant": top2box(cat("content_relevant")),
            "facilitatorKnowledge": top2box(cat("fac_knowledge")),
            "facilitatorEngaging": top2box(cat("fac_engaged")),
        },
        "modality": {
            "virtual": sum(count_norm(ds.get("modality", []), norm_modality, "virtual")
                           for ds in datasets),
            "inPerson": sum(count_norm(ds.get("modality", []), norm_modality, "in person")
                            for ds in datasets),
        },
    }

    # Manager expectations: % NO among answers we can actually read. Anything
    # unrecognised is excluded from BOTH halves rather than silently inflating
    # the denominator.
    me_values = [norm_yes_no(v) for ds in datasets
                 for v in ds.get("manager_exp", []) if v not in (None, "")]
    me_values = [v for v in me_values if v]
    if me_values:
        no_count = sum(1 for v in me_values if v == "no")
        view["noManagerExpectationsPct"] = round(no_count / len(me_values) * 100)
        view["managerExpectationsResponses"] = len(me_values)

    # Same program, split by whether the participant's manager set
    # expectations. Columns are read row-for-row, so index i is one person.
    with_mgr, without_mgr = [], []
    for ds in datasets:
        answers, scores = ds.get("manager_exp") or [], ds.get("nps") or []
        for i in range(min(len(answers), len(scores))):
            if answers[i] in (None, ""):
                continue
            answer = norm_yes_no(answers[i])
            if not answer:
                continue
            try:
                score = float(scores[i])
            except (TypeError, ValueError):
                continue
            (with_mgr if answer == "yes" else without_mgr).append(score)

    # Below ~20 a side, NPS swings wildly on one person — Private L2 in 2026
    # shows a 19-point "effect" off 16 responses. Don't publish those.
    MIN_PER_SIDE = 20
    if len(with_mgr) >= MIN_PER_SIDE and len(without_mgr) >= MIN_PER_SIDE:
        view["managerSplit"] = {
            "withExpectations": {"nps": nps(with_mgr), "n": len(with_mgr)},
            "without": {"nps": nps(without_mgr), "n": len(without_mgr)},
            "gap": nps(with_mgr) - nps(without_mgr),
        }

    return view


def ltf_view(label, datasets):
    """LTF's own view. Its seven statements don't map onto the standard six
    tiles, so it reports them directly: agreement and strength of agreement
    per statement, plus the advocacy question as the headline companion to NPS."""
    def cat(key):
        out = []
        for ds in datasets:
            out.extend(ds.get(key, []))
        return out

    def share(nums, floor):
        return round(sum(1 for v in nums if v >= floor) / len(nums) * 100) if nums else 0

    sessions_all = [s for s in cat("session") if s]
    nps_all = cat("nps")
    participants = len(numeric(nps_all))
    if participants == 0:
        participants = sum(sum(1 for v in ds.get("modality", []) if v) for ds in datasets)

    statements, all_ratings = [], []
    for key, text in LTF_STATEMENTS:
        nums = numeric(cat(key))
        if not nums:
            continue
        all_ratings.extend(nums)
        statements.append({
            "text": text,
            "n": len(nums),
            "top2": share(nums, 4),
            "strongly": share(nums, 5),
            "mean": round(sum(nums) / len(nums), 2),
        })

    advocacy = numeric(cat("recommend_team"))

    return {
        "label": label,
        "type": "ltf",
        "nps": nps(nps_all),
        "participants": participants,
        "sessions": distinct_count(sessions_all),
        "clients": distinct_count(sessions_all),
        "advocacy": share(advocacy, 4),
        "advocacyStrongly": share(advocacy, 5),
        "stronglyAgreeOverall": share(all_ratings, 5),
        "ratingsCount": len(all_ratings),
        "unanimous": sum(1 for s in statements if s["top2"] == 100),
        "statements": statements,
        "modality": {
            "virtual": sum(count_text(ds.get("modality", []), "Virtual") for ds in datasets),
            "inPerson": sum(count_text(ds.get("modality", []), "In person") for ds in datasets),
        },
    }


def learner_view(ws):
    """The 90-day impact view: what people still do, banded by job level."""
    from openpyxl.utils import get_column_letter

    hdr = {get_column_letter(c): " ".join(str(ws.cell(row=1, column=c).value or "").split())
           for c in range(1, ws.max_column + 1)}
    letters = list(hdr)

    def find(needle):
        return next((L for L in letters if needle in hdr[L].lower()), None)

    rows = [r for r in range(2, ws.max_row + 1)
            if any(ws.cell(row=r, column=c).value not in (None, "")
                   for c in range(1, 12))]

    single = {k: find(n) for k, n in LEARNER_SINGLES.items()}
    def cell(letter, r):
        return ws[f"{letter}{r}"].value if letter else None

    # Which band each respondent falls in.
    def band_of(r):
        lvl = str(cell(single["job_level"], r) or "").strip().lower()
        for name, members in LEARNER_BANDS:
            if lvl in members:
                return name
        return None

    # Before/after pairs, keyed by the behavior they measure.
    pairs = {}
    for L in letters:
        h = hdr[L]
        if " - Before EQ Training" in h or " - After EQ Training" in h:
            stem, phase = h.rsplit(" - ", 1)
            pairs.setdefault(stem.rstrip(" ."), {})[phase.split()[0]] = L

    def score(value, reversed_item):
        n = LEARNER_SCALE.get(str(value).strip().lower())
        if n is None:
            return None
        return (6 - n) if reversed_item else n   # higher always means better

    # Multi-select blocks, bounded by hand.
    def block_counts(first_needle, last_needle, subset):
        start = find(first_needle)
        if not start:
            return []
        end = next((L for L in letters[letters.index(start):]
                    if hdr[L].lower() == last_needle), None)
        span = letters[letters.index(start): letters.index(end) + 1 if end else None]
        out = []
        for L in span:
            picked = [r for r in subset if cell(L, r) not in (None, "")]
            if not picked:
                continue
            # The first column's header is the question; its answer text is the option.
            name = hdr[L]
            if L == start:
                name = str(cell(L, picked[0])).strip()
            if name.lower().startswith("other"):
                continue
            out.append({"label": name, "count": len(picked)})
        return sorted(out, key=lambda d: -d["count"])

    def distribution(letter, subset):
        vals = [str(cell(letter, r)).strip() for r in subset
                if cell(letter, r) not in (None, "")]
        total = len(vals)
        return [{"label": k, "count": n, "pct": round(n / total * 100)}
                for k, n in Counter(vals).most_common()] if total else []

    def build(subset):
        behaviors = []
        improved_any, paired_people = set(), set()
        for stem, ph in pairs.items():
            if "Before" not in ph or "After" not in ph:
                continue
            rev = any(k in stem.lower() for k in LEARNER_REVERSED)
            before, after, up = [], [], 0
            for r in subset:
                sb = score(cell(ph["Before"], r), rev)
                sa = score(cell(ph["After"], r), rev)
                if sb is None or sa is None:
                    continue
                before.append(sb); after.append(sa); paired_people.add(r)
                if sa > sb:
                    up += 1; improved_any.add(r)
            if not before:
                continue
            mb, ma = sum(before) / len(before), sum(after) / len(after)
            behaviors.append({
                "text": stem, "reversed": rev, "n": len(before),
                "before": round(mb, 2), "after": round(ma, 2),
                "change": round(ma - mb, 2),
                "improvedPct": round(up / len(before) * 100),
            })
        behaviors.sort(key=lambda b: -b["change"])

        nps_scores = numeric([cell(single["nps"], r) for r in subset])
        applies = distribution(single["applies"], subset)
        often = sum(d["count"] for d in applies if d["label"].lower() in ("always", "usually"))
        applies_total = sum(d["count"] for d in applies)

        return {
            "respondents": len(subset),
            "pairedRespondents": len(paired_people),
            "nps": nps(nps_scores) if nps_scores else None,
            "npsResponses": len(nps_scores),
            "beforeAvg": round(sum(b["before"] for b in behaviors) / len(behaviors), 2)
                         if behaviors else 0,
            "afterAvg": round(sum(b["after"] for b in behaviors) / len(behaviors), 2)
                        if behaviors else 0,
            "improvedAnyPct": round(len(improved_any) / len(paired_people) * 100)
                              if paired_people else 0,
            "appliesOftenPct": round(often / applies_total * 100) if applies_total else 0,
            "behaviors": behaviors,
            "applies": applies,
            "challenges": distribution(single["challenges"], subset),
            "managerReinforcement": distribution(single["manager"], subset),
            "overall": distribution(single["overall"], subset),
            "eiAttributed": avg([cell(single["ei_pct"], r) for r in subset]),
            "blocks": [{"label": lbl, "items": block_counts(a, b, subset)}
                       for lbl, a, b in LEARNER_BLOCKS],
        }

    by_band = {"All": build(rows)}
    for name, _ in LEARNER_BANDS:
        members = [r for r in rows if band_of(r) == name]
        if members:
            by_band[name] = build(members)

    cohorts = sorted({str(cell(single["session"], r)).strip()
                      for r in rows if cell(single["session"], r)})

    return {
        "label": "90-Day Impact",
        "type": "learner",
        "bands": list(by_band),
        "cohorts": len(cohorts),
        "byBand": by_band,
    }


def refresher_view(ws):
    before = col_values(ws, "L")
    after = col_values(ws, "M")
    value_rating = col_values(ws, "AB")
    sessions = col_values(ws, "A")

    def to_num(v):
        if v is None: return None
        return CONFIDENCE_MAP.get(str(v).strip().lower())

    before_n = [n for n in (to_num(v) for v in before) if n is not None]
    after_n = [n for n in (to_num(v) for v in after) if n is not None]

    paired = [(to_num(b), to_num(a)) for b, a in zip(before, after)]
    paired = [(b, a) for b, a in paired if b is not None and a is not None]
    moved_up = sum(1 for b, a in paired if a > b)

    valuable = [v for v in value_rating if v]
    valuable_top = sum(1 for v in valuable if str(v).strip().lower() in VALUABLE_TOP)

    return {
        "label": "Refresher",
        "type": "refresher",
        "participants": len([v for v in sessions if v]),
        "sessions": distinct_count(sessions),
        "confidenceBefore": round(sum(before_n) / len(before_n), 2) if before_n else 0,
        "confidenceAfter": round(sum(after_n) / len(after_n), 2) if after_n else 0,
        "confidenceGrowth": round((sum(after_n)/len(after_n)) - (sum(before_n)/len(before_n)), 2) if before_n and after_n else 0,
        "pctMovedUpInConfidence": round(moved_up / len(paired) * 100) if paired else 0,
        "pctRatedValuable": round(valuable_top / len(valuable) * 100) if valuable else 0,
        "confidenceScale": "1=Slightly · 2=Moderately · 3=Confident · 4=Fully",
    }


POSITIVE_WORDS = {
    "amazing", "appreciate", "appreciated", "authentic", "awesome", "beneficial",
    "best", "brilliant", "captivating", "dynamic", "effective", "empowered",
    "empowering", "engaging", "enjoyed", "enlightening", "excellent", "exceptional",
    "fabulous", "fantastic", "genuine", "good", "grateful", "great", "happy",
    "helpful", "highly", "impactful", "incredible", "inspired", "inspiring",
    "insightful", "invaluable", "love", "loved", "magnificent", "meaningful",
    "memorable", "motivated", "outstanding", "passionate", "perfect", "phenomenal",
    "powerful", "profound", "recommend", "rich", "thank", "thoughtful", "thrilled",
    "transformative", "valuable", "wonderful",
    # phrases will be matched separately below
}
POSITIVE_PHRASES = [
    "eye opening", "eye-opening", "must take", "must do", "must attend",
    "best in", "highly recommend", "well worth", "go for it", "top notch",
    "in the business",
]
EXCLUDE_PHRASES = [
    "log off", "block diary", "ask organiser", "ask organizer",
    "internal milestones", "should have been", "would have been",
    "needs to be improved", "n/a", "na ", "none ", "no comment",
    "wish there had been", "missed the mark", "didn't enjoy", "did not enjoy",
    # Advisory / cautionary tone — not testimonial material
    "get lost", "don't get lost", "dont get lost",
    "pay attention", "pay good attention",
    "be ready", "you'll need", "make sure you",
    "so you don't", "so you dont",
]

def quote_score(text):
    t = " " + text.lower() + " "
    if any(p in t for p in EXCLUDE_PHRASES):
        return -99
    pos = sum(1 for w in POSITIVE_WORDS if f" {w} " in t or f" {w}." in t or f" {w}!" in t or f" {w}," in t)
    pos += sum(2 for p in POSITIVE_PHRASES if p in t)
    return pos


def substance_bonus(text):
    """Reward thoughtful, specific testimonials so they don't get buried under
    short buzzword-dense blurbs. Rewards multi-sentence depth, length, and
    first-person reflection — the marks of a credible, quotable story."""
    low = " " + text.lower() + " "
    sentences = sum(text.count(c) for c in ".!?")
    bonus = min(sentences, 4)            # up to +4 for multi-sentence depth
    if len(text) >= 150: bonus += 1
    if len(text) >= 280: bonus += 2      # a substantial, fleshed-out reflection
    if " i " in low or low.lstrip().startswith("i "):
        bonus += 1                       # personal, first-person voice
    return bonus


def rank_value(text, score, names_trainer):
    """Sort key for testimonials: positivity + substance + a nudge for quotes
    that name the trainer. Higher ranks first."""
    return score + substance_bonus(text) + (1 if names_trainer else 0)


def mentions_facilitator(text, facilitator):
    """True if the quote text mentions any token of the facilitator's name (>=3 chars)."""
    if not facilitator:
        return False
    t = text.lower()
    tokens = [tok for tok in facilitator.lower().split() if len(tok) >= 3]
    return any(tok in t for tok in tokens)


def company_from_email(email):
    """Derive a short company tag from an email's domain, e.g. jane@syf.com -> 'SYF'.
    Returns '' for blank/personal/generic domains (gmail, yahoo, ...)."""
    if not email or "@" not in str(email):
        return ""
    domain = str(email).split("@")[-1].strip().lower().strip(".")
    if "." not in domain:
        return ""
    label = domain.split(".")[0]
    if not label or label in GENERIC_EMAIL_DOMAINS:
        return ""
    return label.upper()


def _clean_name_token(s):
    """Keep letters/hyphens/apostrophes; drop stray punctuation and mojibake."""
    return "".join(ch for ch in s if ch.isalpha() or ch in "-'").strip("-'")


def display_name(first, last, name_optional):
    """Build the display name per NAME_STYLE. Falls back to the free-text
    'Name (optional)' field when First/Last aren't filled in. '' if unknown."""
    first = str(first).strip() if first else ""
    last = str(last).strip() if last else ""
    if not first and name_optional:
        parts = str(name_optional).strip().split()
        if parts:
            first = parts[0]
            last = parts[-1] if len(parts) > 1 else ""
    first = _clean_name_token(first)
    last = _clean_name_token(last)
    if not first:
        return ""
    if NAME_STYLE == "full":
        return (first + " " + last).strip()
    if NAME_STYLE == "first_initial" and last:
        return f"{first} {last[0]}."
    return first


def detect_extra_cols(ws):
    """Locate name/email/strengths columns by header text. Returns {key: 'G', ...}."""
    from openpyxl.utils import get_column_letter
    found = {}
    for c in range(1, min(ws.max_column, 60) + 1):
        h = ws.cell(row=1, column=c).value
        if not h:
            continue
        hl = str(h).strip().lower()
        for key, needles in EXTRA_HEADER_NEEDLES.items():
            if key in found:
                continue
            # A real contact column is labelled, not asked: "Email Address".
            # Survey *questions* also contain the word — LTF's opt-in ends
            # "...please add your email address here!" — and matching one would
            # turn a "may we contact you" answer into a public company tag.
            if key == "email" and len(hl) > 40:
                continue
            if any(n in hl for n in needles):
                found[key] = get_column_letter(c)
    return found


def best_quotes(ds, program_label, max_n=None):
    """Collect all quotes that are either clearly positive OR name the trainer.

    Pulls from two free-text sources: "share with future facilitators" and
    "facilitator's strengths". Each qualifying answer becomes a testimonial,
    attributed with the respondent's name (+ company when a corporate email
    is on file). No cap by default; excluded only on an EXCLUDE_PHRASES marker.
    """
    quotes_raw = ds.get("quote", [])
    strengths_raw = ds.get("strengths", [])
    facilitators = ds.get("facilitator", [])
    emails = ds.get("email", [])
    firsts = ds.get("first", [])
    lasts = ds.get("last", [])
    names_opt = ds.get("name_optional", [])

    def at(lst, i):
        return lst[i] if i < len(lst) else None

    n_rows = max((len(x) for x in (quotes_raw, strengths_raw, facilitators,
                                   emails, firsts, lasts, names_opt)), default=0)

    pairs = []
    for i in range(n_rows):
        fac = ""
        f = at(facilitators, i)
        if f:
            fac = str(f).strip()
        name = display_name(at(firsts, i), at(lasts, i), at(names_opt, i))
        company = company_from_email(at(emails, i))

        for src in (at(quotes_raw, i), at(strengths_raw, i)):
            if not src:
                continue
            text = str(src).strip()
            if not (25 <= len(text) <= 500): continue
            score = quote_score(text)
            if score == -99: continue  # killed by exclude phrases

            names_trainer = mentions_facilitator(text, fac)

            # Include if positive OR names the trainer (the two "good ones" criteria)
            if score < 1 and not names_trainer:
                continue

            rank = rank_value(text, score, names_trainer)
            pairs.append((rank, text, fac, names_trainer, name, company))

    # Sort: substance + positivity desc, then longer first as a tie-breaker.
    pairs.sort(key=lambda x: (-x[0], -len(x[1])))
    seen, picked = set(), []
    for _, text, fac, _, name, company in pairs:
        key = text[:60].lower()
        if key in seen: continue
        seen.add(key)
        picked.append({"quote": text, "program": program_label, "facilitator": fac,
                       "name": name, "company": company})
        if max_n is not None and len(picked) >= max_n:
            break
    return picked


# Standard views, in display order. Each names the raw sheet families it draws
# from, so a family can span several source sheets (2026 + its 2025 history).
STANDARD_VIEWS = [
    ("ttt-summary",     "Train the Trainer — Summary", ["TTTL1", "TTTL2", "TTTTeams"]),
    ("ttt-l1",          "Train the Trainer — Level 1", ["TTTL1"]),
    ("ttt-l2",          "Train the Trainer — Level 2", ["TTTL2"]),
    ("ttt-teams",       "Train the Trainer — Teams",   ["TTTTeams"]),
    ("private-summary", "Private Program — Summary",   ["PrivateL1", "PrivateL2"]),
    ("private-l1",      "Private Program — Level 1",   ["PrivateL1"]),
    ("private-l2",      "Private Program — Level 2",   ["PrivateL2"]),
    ("public-l1",       "Public Program",              ["PublicL1"]),
    ("custom-summary",  "Custom Programs — Summary",   ["Custom Programs", "Policing"]),
    ("custom-misc",     "Custom Programs — Misc.",     ["Custom Programs"]),
    ("custom-policing", "Custom Programs — EQ in Policing", ["Policing"]),
]

# What the dashboard shows before anyone touches the year control.
DEFAULT_PERIOD = "2026"


def build_standard_views(raw, period):
    """Every standard view for one period. 'all' spans every family, including
    LTF, which is why it's built from raw rather than STANDARD_VIEWS."""
    out = {}
    everything = [d for lst in raw.values() for d in lst]
    out["all"] = standard_view("All Programs",
                               [filter_period(d, period) for d in everything])
    for key, label, families in STANDARD_VIEWS:
        dsets = [filter_period(d, period)
                 for fam in families for d in raw.get(fam, [])]
        out[key] = standard_view(label, dsets)
    return out


PROGRAM_LABELS = {
    "TTTL1": "TTT L1", "TTTL2": "TTT L2", "TTTTeams": "TTT Teams",
    "PrivateL1": "Private L1", "PrivateL2": "Private L2",
    "PublicL1": "Public L1", "Custom Programs": "Custom Programs",
    "LTF-Private": "LTF — Private", "LTF-Public": "LTF — Public",
    "LTF-TTT": "LTF — TTT", "Policing": "EQ in Policing",
}


def update_trainers_tab(cols_by_sheet):
    """Rewrites the Trainers tab. One OVERALL row per trainer plus a per-program
    drill-down row for each program they've taught.

    Takes the column map resolved in main() rather than re-resolving: this
    function opens the workbook without data_only, where value-shape checks
    would be looking at formulas instead of results.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict

    wb = openpyxl.load_workbook(SRC)
    if "Trainers" in wb.sheetnames:
        del wb["Trainers"]
    insert_idx = 2 if "Quarterly" in wb.sheetnames else (1 if "Summary" in wb.sheetnames else 0)
    ws = wb.create_sheet("Trainers", insert_idx)
    ws.sheet_view.showGridLines = False

    NAVY = "002D61"
    GOLD_LIGHT = "FFF4D6"
    BORDER = "E5E9F0"

    # Title + subtitle
    ws.merge_cells("A1:I1")
    ws["A1"] = "Trainer Performance — Year to Date"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:I2")
    ws["A2"] = ("OVERALL row = trainer's combined performance across all programs they've taught. "
                "Drill-down rows beneath show the same trainer per program. "
                "Refresh by running update-data.bat after adding survey rows.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = ["Trainer", "Program / Scope", "Sessions", "Respondents",
               "NPS", "% Apply on Job", "% Engaging",
               "% Virtual", "% In-Person"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 38

    widths = [26, 22, 11, 13, 9, 14, 14, 12, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Gather per-trainer × per-program data
    # trainer_data[trainer] = {program_label: {nps, apply, engage, virtual, inperson, sessions}}
    trainer_data = defaultdict(lambda: defaultdict(lambda: {
        "nps": [], "apply": [], "engage": [], "modality": [], "sessions": set()
    }))
    for sheet_name, cols in cols_by_sheet.items():
        if sheet_name not in wb.sheetnames or not cols.get("facilitator"):
            continue
        raw_ws = wb[sheet_name]
        program_label = PROGRAM_LABELS.get(sheet_name, sheet_name)
        for r in range(2, raw_ws.max_row + 1):
            fac = raw_ws[f"{cols['facilitator']}{r}"].value
            if not fac: continue
            fac = str(fac).strip()
            if not fac: continue
            g = trainer_data[fac][program_label]
            g["nps"].append(raw_ws[f"{cols['nps']}{r}"].value)
            g["apply"].append(raw_ws[f"{cols['apply_on_job']}{r}"].value)
            g["engage"].append(raw_ws[f"{cols['fac_engaged']}{r}"].value)
            mod = raw_ws[f"{cols['modality']}{r}"].value
            if mod: g["modality"].append(str(mod).strip())
            sess = raw_ws[f"{cols['session']}{r}"].value
            if sess: g["sessions"].add(str(sess).strip())

    def aggregate(groups_dict):
        """Combine all programs for a trainer into one row of stats."""
        agg = {"nps": [], "apply": [], "engage": [], "modality": [], "sessions": set()}
        for prog, g in groups_dict.items():
            agg["nps"].extend(g["nps"])
            agg["apply"].extend(g["apply"])
            agg["engage"].extend(g["engage"])
            agg["modality"].extend(g["modality"])
            agg["sessions"] |= g["sessions"]
        return agg

    def stats_row(label_program, g):
        n_resp = len(numeric(g["nps"]))
        if n_resp == 0:
            return None
        n_virtual = sum(1 for m in g["modality"] if m.lower() == "virtual")
        n_inperson = sum(1 for m in g["modality"] if m.lower() == "in person")
        n_total_mod = n_virtual + n_inperson
        v_pct = round(n_virtual / n_total_mod * 100) if n_total_mod else 0
        i_pct = round(n_inperson / n_total_mod * 100) if n_total_mod else 0
        return [
            label_program,
            len(g["sessions"]),
            n_resp,
            nps(g["nps"]),
            f'{top2box(g["apply"])}%',
            f'{top2box(g["engage"])}%',
            f'{v_pct}%' if n_total_mod else '-',
            f'{i_pct}%' if n_total_mod else '-',
        ]

    # Render: trainer (alpha order), OVERALL row, then per-program rows beneath.
    excel_row = 5
    bold_navy = Font(bold=True, color=NAVY, size=11)
    bold = Font(bold=True, size=11)
    light = Font(color="555555", size=11)
    overall_fill = PatternFill("solid", fgColor=GOLD_LIGHT)
    thin = Side(border_style="thin", color=BORDER)
    bot = Border(bottom=Side(border_style="thin", color="CCCCCC"))

    for trainer in sorted(trainer_data.keys(), key=str.lower):
        programs = trainer_data[trainer]
        # OVERALL row
        agg = aggregate(programs)
        row_vals = stats_row("OVERALL", agg)
        if not row_vals: continue
        ws.cell(row=excel_row, column=1, value=trainer).font = bold_navy
        for j, v in enumerate(row_vals, start=2):
            c = ws.cell(row=excel_row, column=j, value=v)
            c.alignment = Alignment(horizontal="center" if j > 2 else "left", vertical="center")
            c.font = bold
            c.fill = overall_fill
        ws.cell(row=excel_row, column=1).fill = overall_fill
        excel_row += 1

        # Per-program rows
        for prog in sorted(programs.keys()):
            row_vals = stats_row(prog, programs[prog])
            if not row_vals: continue
            ws.cell(row=excel_row, column=1, value="").alignment = Alignment(horizontal="left", indent=1)
            for j, v in enumerate(row_vals, start=2):
                c = ws.cell(row=excel_row, column=j, value=v)
                c.alignment = Alignment(horizontal="center" if j > 2 else "left", vertical="center", indent=1 if j == 2 else 0)
                c.font = light
            excel_row += 1
        # blank separator row
        excel_row += 1

    ws.freeze_panes = "A5"
    wb.save(SRC)
    print(f"Trainers tab updated: {len(trainer_data)} trainers")


def session_rows(raw):
    """One row per session, for the Excel-only Sessions tab.

    Session names carry client names, so this deliberately never reaches
    data.json — it stays in the workbook on OneDrive.
    """
    from collections import defaultdict
    acc = defaultdict(lambda: {
        "dates": [], "nps": [], "apply": [], "engage": [],
        "modality": [], "facilitators": Counter(), "program": ""})

    for family, datasets in raw.items():
        label = PROGRAM_LABELS.get(family, family)
        for ds in datasets:
            dates = ds.get("_date", [])
            def at(key, i):
                col = ds.get(key) or []
                return col[i] if i < len(col) else None
            for i in range(len(dates)):
                sess = at("session", i)
                if not sess:
                    continue
                g = acc[str(sess).strip()]
                g["program"] = label
                if dates[i]:
                    g["dates"].append(dates[i])
                for key, bucket in (("nps", "nps"), ("apply_on_job", "apply"),
                                    ("fac_engaged", "engage")):
                    v = at(key, i)
                    if v not in (None, ""):
                        g[bucket].append(v)
                m = norm_modality(at("modality", i)) if at("modality", i) else ""
                if m:
                    g["modality"].append(m)
                f = at("facilitator", i)
                if f:
                    g["facilitators"][str(f).strip()] += 1

    rows = []
    for sess, g in acc.items():
        scores = numeric(g["nps"])
        if not scores:
            continue
        when = min(g["dates"]) if g["dates"] else None
        per = period_of(when) or {}
        virtual = sum(1 for m in g["modality"] if m == "virtual")
        in_person = sum(1 for m in g["modality"] if m == "in person")
        rows.append({
            "year": per.get("year", ""),
            "quarter": per.get("quarter", ""),
            "date": when.date().isoformat() if when else "",
            "program": g["program"],
            "session": sess,
            "facilitator": (g["facilitators"].most_common(1)[0][0]
                            if g["facilitators"] else ""),
            "responses": len(scores),
            "nps": nps(scores),
            "promoters": sum(1 for s in scores if s >= 9),
            "passives": sum(1 for s in scores if 7 <= s <= 8),
            "detractors": sum(1 for s in scores if s <= 6),
            "apply": top2box(g["apply"]) if g["apply"] else "",
            "engaging": top2box(g["engage"]) if g["engage"] else "",
            "modality": ("Virtual" if virtual and not in_person else
                         "In person" if in_person and not virtual else
                         "Mixed" if virtual and in_person else ""),
        })
    rows.sort(key=lambda r: (str(r["date"]), r["session"]), reverse=True)
    return rows


def update_sessions_tab(rows):
    """Writes the Sessions tab: every session, newest first, with year and
    quarter as their own filterable columns."""
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.load_workbook(SRC)
    if "Sessions" in wb.sheetnames:
        del wb["Sessions"]
    idx = (wb.sheetnames.index("Trainers") + 1) if "Trainers" in wb.sheetnames else 0
    ws = wb.create_sheet("Sessions", idx)
    ws.sheet_view.showGridLines = False

    NAVY, LOW, MID, HIGH = "002D61", "FDE7E9", "FFF4D6", "E4F7EF"

    ws.merge_cells("A1:N1")
    ws["A1"] = "Every Session — Internal View"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:N2")
    ws["A2"] = ("Session names include client names — this tab is deliberately NOT published "
                "to the dashboard. Use the filter arrows on row 4 to slice by Year or Quarter. "
                "NPS is shaded: red below 50, amber 50-69, green 70+.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = ["Year", "Quarter", "Date", "Program", "Session", "Facilitator",
               "Responses", "NPS", "Promoters", "Passives", "Detractors",
               "% Apply on Job", "% Engaging", "Modality"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 34

    for i, w in enumerate([8, 10, 12, 16, 54, 20, 11, 8, 11, 10, 11, 14, 12, 11], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    for r, row in enumerate(rows, start=5):
        vals = [row["year"], row["quarter"], row["date"], row["program"],
                row["session"], row["facilitator"], row["responses"], row["nps"],
                row["promoters"], row["passives"], row["detractors"],
                (row["apply"] / 100 if row["apply"] != "" else ""),
                (row["engaging"] / 100 if row["engaging"] != "" else ""),
                row["modality"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = Alignment(horizontal="left" if j in (4, 5, 6) else "center",
                                    vertical="center")
            if j in (12, 13) and v != "":
                c.number_format = "0%"
        nps_cell = ws.cell(row=r, column=8)
        nps_cell.font = Font(bold=True)
        nps_cell.fill = PatternFill("solid", fgColor=(
            LOW if row["nps"] < 50 else MID if row["nps"] < 70 else HIGH))

    ws.auto_filter.ref = f"A4:N{4 + len(rows)}"
    ws.freeze_panes = "A5"
    wb.save(SRC)
    print(f"Sessions tab updated: {len(rows)} sessions")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    ltf_present = ltf_sheets(wb.sheetnames)

    col_warnings = []
    cols_by_sheet = all_sheet_cols(wb, warn=col_warnings)
    for w in col_warnings:
        print(f"  ! {w}")

    raw = {}
    for sheet_name, cols in cols_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            print(f"  skipping missing sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        ds = collect_sheet(ws, cols)
        # Attribution + extra quote source, located by header text.
        extra = detect_extra_cols(ws)
        for key in ("email", "first", "last", "name_optional", "strengths"):
            ds[key] = col_values(ws, extra[key]) if extra.get(key) else []
        raw[sheet_name] = [ds]

    # 2025 history, from its own workbook, appended to the family it continues.
    wb_learner = None
    if LEGACY_SRC.exists():
        wb_old = load_readonly(LEGACY_SRC, data_only=True)
        wb_learner = wb_old
        for sheet_old, family in LEGACY_SHEETS.items():
            if sheet_old not in wb_old.sheetnames:
                print(f"  skipping missing 2025 sheet: {sheet_old}")
                continue
            ws_old = wb_old[sheet_old]
            cols_old = resolve_cols(ws_old, sheet_old, warn=col_warnings)
            ds_old = collect_sheet(ws_old, cols_old)
            extra = detect_extra_cols(ws_old)
            for key in ("email", "first", "last", "name_optional", "strengths"):
                ds_old[key] = col_values(ws_old, extra[key]) if extra.get(key) else []
            raw.setdefault(family, []).append(ds_old)
            dated = sum(1 for d in ds_old["_date"] if d)
            print(f"  + 2025 {sheet_old} -> {family}: {dated} dated responses")
    else:
        print(f"  ! 2025 workbook not found at {LEGACY_SRC}")

    # LTF again, this time under its own question names, for its own view.
    ltf_raw = {name: collect_sheet(wb[name], LTF_COLS) for name in ltf_present}

    views = {}

    # Standard views for the default period. The year control swaps these out
    # for the matching entry in viewsByPeriod.
    views.update(build_standard_views(raw, DEFAULT_PERIOD))

    # Refresher is its own instrument on its own timeline — not year-filtered.
    views["refresher"] = refresher_view(wb["Refresher"])

    # 90-day impact survey — its own instrument too, banded by job level.
    if wb_learner is not None and LEARNER_SHEET in wb_learner.sheetnames:
        views["learner"] = learner_view(wb_learner[LEARNER_SHEET])

    # Leading Through Friction — summary plus one view per delivery mode that
    # actually has data. Views the dashboard doesn't find, it hides.
    if ltf_present:
        views["ltf-summary"] = ltf_view(
            "Leading Through Friction", [ltf_raw[s] for s in ltf_present])
        for s in ltf_present:
            key, delivery = LTF_DELIVERIES[s.split("-", 1)[1]]
            views[key] = ltf_view(
                f"Leading Through Friction — {delivery}", [ltf_raw[s]])

    # Quotes
    testimonials = []
    label_for = {
        "TTTL1": "Train the Trainer L1", "TTTL2": "Train the Trainer L2",
        "TTTTeams": "TTT for Teams", "PrivateL1": "Private Program L1",
        "PrivateL2": "Private Program L2", "PublicL1": "Public Program",
        "Custom Programs": "Custom Program", "Policing": "EQ in Policing",
    }
    view_for = {
        "TTTL1": "ttt-l1", "TTTL2": "ttt-l2", "TTTTeams": "ttt-teams",
        "PrivateL1": "private-l1", "PrivateL2": "private-l2",
        "PublicL1": "public-l1", "Custom Programs": "custom-misc",
        "Policing": "custom-policing",
    }
    for s in ltf_present:
        key, delivery = LTF_DELIVERIES[s.split("-", 1)[1]]
        label_for[s] = f"Leading Through Friction ({delivery})"
        view_for[s] = key
    for sheet_name, datasets in raw.items():
        for ds in datasets:
            for period in periods_in([ds]) or [DEFAULT_PERIOD]:
                for q in best_quotes(filter_period(ds, period), label_for[sheet_name]):
                    q["view"] = view_for[sheet_name]
                    q["period"] = period
                    testimonials.append(q)

    all_periods = periods_in([d for lst in raw.values() for d in lst])
    views_by_period = {p: build_standard_views(raw, p)
                       for p in ["all"] + all_periods}

    payload = {
        "meta": {
            "lastUpdated": datetime.now().strftime("%Y-%m-%d"),
            "reportingPeriod": "Q1 2026 to date",
            "periods": all_periods,
            "defaultPeriod": DEFAULT_PERIOD,
        },
        "views": views,
        "viewsByPeriod": views_by_period,
        "testimonials": testimonials,
    }

    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"\nWrote: {OUT}")

    # Refresh the internal-only workbook tabs
    update_trainers_tab(cols_by_sheet)
    update_sessions_tab(session_rows(raw))
    print("\nQuick summary of computed views:")
    for k, v in views.items():
        if v.get("type") == "refresher":
            print(f"  {k:18s} participants={v['participants']:>3}  growth={v['confidenceGrowth']}  valuable%={v['pctRatedValuable']}")
        elif v.get("type") == "learner":
            a = v["byBand"]["All"]
            print(f"  {k:18s} n={a['respondents']:>3}  paired={a['pairedRespondents']:>3}  "
                  f"{a['beforeAvg']}->{a['afterAvg']}  NPS@90d={a['nps']}  "
                  f"applying={a['appliesOftenPct']}%  bands={len(v['bands'])}")
        elif v.get("type") == "ltf":
            print(f"  {k:18s} NPS={v['nps']:>3}  N={v['participants']:>3}  "
                  f"advocacy={v['advocacy']}%  unanimous={v['unanimous']}/{len(v['statements'])}")
        else:
            extra = f"  noManager%={v.get('noManagerExpectationsPct','-')}" if "noManagerExpectationsPct" in v else ""
            print(f"  {k:18s} NPS={v['nps']:>3}  N={v['participants']:>3}  applyOnJob={v['topBox']['applyOnJob']}%{extra}")


if __name__ == "__main__":
    main()
